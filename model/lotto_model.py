"""
Initialisierung model/lotto_model.py …
Ziel:

- Modell LottoZiehung
- Speichern in Excel + JSON
- Duplikatprüfung
- Ladefunktion für bestehende Daten
"""

import os
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# Basispfade (relativ zum Projektordner)
DATA_DIR = Path(__file__).resolve().parent.parent / "data"
EXCEL_PATH = DATA_DIR / "ziehungen.xlsx"
JSON_PATH = DATA_DIR / "ziehungen.json"


class LottoZiehung:
    """
    Repräsentiert eine einzelne Lotto-Ziehung.
    Enthält Datum, gezogene Zahlen, Superzahl und Quoten.
    """
    def __init__(self, datum: str, zahlen: list[int], superzahl: int, quoten: dict):
        self.datum = datum  # Format: "YYYY-MM-DD"
        self.zahlen = zahlen  # Liste von 6 Lottozahlen
        self.superzahl = superzahl  # Superzahl als Zahl
        self.quoten = quoten  # Gewinnquoten als Dictionary: {Klasse: Gewinn}

    def to_dict(self):
        """
        Wandelt die Ziehung in ein serialisierbares Dictionary um.
        Wird für JSON-Speicherung verwendet.
        """
        return {
            "datum": self.datum,
            "zahlen": self.zahlen,
            "superzahl": self.superzahl,
            "quoten": self.quoten,
        }


def lade_bestehende_ziehungen():
    """
    Lädt bereits gespeicherte Ziehungen aus der JSON-Datei.
    Gibt eine Liste von Dictionaries zurück.
    """
    if not JSON_PATH.exists():
        return []  # Noch keine Daten vorhanden
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def ist_neue_ziehung(ziehung: LottoZiehung, bestehende: list[dict]) -> bool:
    """
    Prüft, ob die Ziehung bereits in der bestehenden Liste enthalten ist.
    Rückgabe: True, wenn die Ziehung NEU ist.
    """
    return all(z["datum"] != ziehung.datum for z in bestehende)


def speichere_ziehung(ziehung: LottoZiehung):
    """
    Speichert eine neue Ziehung in JSON + Excel, falls sie noch nicht vorhanden ist.
    """
    # Sicherstellen, dass der data/ Ordner existiert
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # JSON-Aktualisierung
    daten = lade_bestehende_ziehungen()
    if ist_neue_ziehung(ziehung, daten):
        daten.append(ziehung.to_dict())
        with open(JSON_PATH, "w", encoding="utf-8") as f:
            json.dump(daten, f, indent=2, ensure_ascii=False)
        print(f"[✓] JSON aktualisiert: {ziehung.datum}")
    else:
        print(f"[!] Ziehung {ziehung.datum} bereits vorhanden (JSON)")

    #  Excel-Aktualisierung
    if not EXCEL_PATH.exists():
        # Neue Excel-Datei mit Kopfzeile erzeugen
        wb = Workbook()
        ws = wb.active
        ws.title = "Ziehungen"
        ws.append(["Datum", "Zahlen", "Superzahl", "Quoten (JSON)"])
    else:
        # Bestehende Datei laden
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

    # Bereits gespeicherte Daten prüfen (Datum in Spalte A)
    existing_dates = [str(cell.value) for cell in ws["A"] if cell.value]
    if ziehung.datum not in existing_dates:
        ws.append([
            ziehung.datum,
            ", ".join(map(str, ziehung.zahlen)),  # Zahlen als Komma-getrennte Zeichenkette
            ziehung.superzahl,
            json.dumps(ziehung.quoten, ensure_ascii=False)  # Quoten als JSON-Text
        ])
        wb.save(EXCEL_PATH)
        print(f"[✓] Excel aktualisiert: {ziehung.datum}")
    else:
        print(f"[!] Ziehung {ziehung.datum} bereits vorhanden (Excel)")
